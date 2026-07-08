#!/usr/bin/env python3
"""v5 §3-2 — build the gold review workbook with the five anti-circularity rules.

  Rule 1  gold = only human-confirmed cells (확정=Y). Drafts are seeds, not gold.
  Rule 2  anti-anchoring: the review sheet shows the SOURCE TEXT as primary
          material; method outputs (룰/LLM/하이브리드) are NOT shown at all.
          Drafts are generated fresh from the raw text, never copied from methods.
  Rule 3  edit rate: drafts live in their own column; the 정답 column starts
          EMPTY (no prefill!), so draft-vs-final edits are measurable.
  Rule 4  blind subset: ~15% of docs get NO visible draft; their hidden drafts
          are stored for the later anchoring check (13_gold_ingest.py).
  Rule 5  time/date gold at raw + date_kind level — the notation asks for
          `제목 | 타입 | 날짜(raw ok: Week N) | 날짜종류`, never resolved dates.

Outputs (all git-ignored, data/gold/):
  gold_review.xlsx   검수(리뷰 입력) + 원문(소스) + 안내(rules/notation) sheets
  drafts.jsonl       ALL drafts incl. hidden blind ones (for the ingest metrics)

Usage:
  # batch 1 — the 37 Excel syllabi:
  python scripts/12_gold_review_build.py [--xlsx ParserTest.xlsx] [--blind 0.15]
  # batch 2+ — expand gold over the normalized corpus (v5 §4-3: 37 is a start):
  python scripts/12_gold_review_build.py --source corpus --n 40 --out-prefix batch2
"""
from __future__ import annotations

import argparse
import concurrent.futures as cf
import json
import random
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from syllabus_classifier.common.env import load_env_key
from syllabus_classifier.eval.excel_harness import FIELDS, load_rows

NOTATION = (
    "수업시간: `Mon 11:00-11:50 ; Wed 11:00-11:50` (요일 Mon~Sun, 24h)\n"
    "이벤트: `제목 | 타입(exam/assignment/other) | 날짜 | 날짜종류` 를 ` ; `로 연결.\n"
    "  날짜는 원문 수준(raw): `YYYY-MM-DD` 또는 `Week N` / `Week 8 Thu` / `시작~끝`.\n"
    "  날짜종류: absolute / relative / uncertain / recurring.  실제 날짜로 변환 금지(Rule 5).\n"
    "무기한과제: 마감 없는 과제 제목만 ` ; `로.\n"
    "주차별내용: `Week N: 내용` 을 ` ; `로 (내용은 원문 언어).\n"
    "연락처: `email ; phone` (있는 것만, 이메일 우선).\n"
    "대학=학교(학과 아님), 학년도=학사연도(인쇄/출력일 아님), 학기 ∈ {1,2,여름,겨울}.\n"
    "값이 원문에 없으면 반드시 빈칸 — 지어내지 말 것."
)

_SYSTEM = (
    "You draft gold labels for a syllabus-parsing evaluation. Extract the fields "
    "from the raw syllabus text EXACTLY as evidenced — no invention. If a field "
    "is not in the text, return null. Dates stay RAW (YYYY-MM-DD or 'Week N' or "
    "'Week N Thu' or 'start~end') with 날짜종류 in {absolute, relative, uncertain, "
    "recurring} — NEVER convert week references to real dates. Notation:\n"
    + NOTATION +
    "\nReturn a JSON object with exactly these keys: " + ", ".join(FIELDS)
)


def _sim(a: dict, b: dict) -> float:
    """Field-value overlap between two drafts over the union of non-empty fields."""
    import re as _re

    def norm(v):
        return _re.sub(r"\s+", " ", str(v)).strip().lower() if v else None

    keys = [f for f in FIELDS if norm(a.get(f)) or norm(b.get(f))]
    if not keys:
        return 0.0
    same = sum(1 for f in keys if norm(a.get(f)) == norm(b.get(f)))
    return same / len(keys)


def _cluster_by_draft_similarity(rows: list[dict], drafts: dict, threshold: float = 0.8) -> list[list[str]]:
    """Greedy union of docs whose drafts overlap >= threshold (near-duplicates)."""
    sids = [r["syllabus_id"] for r in rows]
    parent = {s: s for s in sids}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    for i, a in enumerate(sids):
        for b in sids[i + 1:]:
            if _sim(drafts.get(a, {}), drafts.get(b, {})) >= threshold:
                parent[find(a)] = find(b)
    groups: dict[str, list[str]] = {}
    for s in sids:
        groups.setdefault(find(s), []).append(s)
    return sorted(groups.values(), key=lambda g: (len(g), g[0]))


def draft_one(client, model: str, row: dict) -> dict:
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "system", "content": _SYSTEM},
                  {"role": "user", "content": row["source_text"][:24000]}],
        response_format={"type": "json_object"},
        temperature=0,
        timeout=60,
    )
    data = json.loads(resp.choices[0].message.content)
    return {f: (str(data[f]).strip() if data.get(f) not in (None, "", []) else None) for f in FIELDS}


def corpus_rows(n: int, seed: int, exclude_doc_ids: set = frozenset(), id_prefix: str = "B2") -> list[dict]:
    """Sample normalized corpus docs (stratified by school folder) as review rows.
    Skips docs with no usable text (needs_ocr/failed). syllabus_id = B2-### with
    the real doc_id kept alongside for later joins."""
    import json as _json
    from collections import defaultdict

    from syllabus_classifier.common.config import load_config, resolve_path
    from syllabus_classifier.extract.normalize_doc import NormalizedDoc

    norm_dir = resolve_path(load_config("data.yaml")["paths"]["normalized_dir"])
    by_school = defaultdict(list)
    for fp in sorted(norm_dir.glob("*.json")):
        doc = NormalizedDoc.from_dict(_json.loads(fp.read_text(encoding="utf-8")))
        text = doc.full_text.strip()
        if doc.extraction_quality in ("failed", "needs_ocr") or len(text) < 200:
            continue
        if doc.doc_id in exclude_doc_ids:      # never re-review a prior batch's doc
            continue
        by_school[doc.doc_id.split("__", 1)[0]].append((doc.doc_id, text))

    rng = random.Random(seed)
    per = max(1, -(-n // max(len(by_school), 1)))
    picked = []
    for school in sorted(by_school):
        docs = by_school[school]
        picked.extend(rng.sample(docs, min(per, len(docs))))
    rng.shuffle(picked)
    picked = picked[:n]
    return [{"syllabus_id": f"{id_prefix}-{i+1:03d}", "doc_id": doc_id, "source_text": text}
            for i, (doc_id, text) in enumerate(picked)]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", choices=["excel", "corpus"], default="excel")
    ap.add_argument("--xlsx", default="ParserTest.xlsx")
    ap.add_argument("--n", type=int, default=40, help="corpus docs to sample (--source corpus)")
    ap.add_argument("--out-prefix", default="", help="suffix for output files (e.g. batch2)")
    ap.add_argument("--exclude-drafts", nargs="*", default=[],
                    help="prior drafts jsonl(s); their doc_ids are excluded from sampling")
    ap.add_argument("--blind", type=float, default=0.15)
    ap.add_argument("--model", default="gpt-4o-mini")
    ap.add_argument("--workers", type=int, default=6)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    if not load_env_key():
        print("ERROR: OPENAI_API_KEY not found")
        return 1
    from openai import OpenAI
    client = OpenAI(timeout=60, max_retries=3)

    if args.source == "corpus":
        exclude = set()
        for path in args.exclude_drafts:
            for line in Path(path).read_text(encoding="utf-8").splitlines():
                d = json.loads(line)
                if d.get("doc_id"):
                    exclude.add(d["doc_id"])
        # id prefix from the batch name ("batch3" -> "B3") so batches never collide
        m = re.search(r"(\d+)", args.out_prefix or "")
        id_prefix = f"B{m.group(1)}" if m else "B2"
        rows = corpus_rows(args.n, args.seed, exclude_doc_ids=exclude, id_prefix=id_prefix)
    else:
        rows = [r for r in load_rows(args.xlsx) if r["source_text"]]

    drafts: dict[str, dict] = {}
    with cf.ThreadPoolExecutor(max_workers=args.workers) as ex:
        futs = {ex.submit(draft_one, client, args.model, r): r["syllabus_id"] for r in rows}
        for fut in cf.as_completed(futs):
            sid = futs[fut]
            try:
                drafts[sid] = fut.result()
            except Exception as e:
                print(f"  draft failed for {sid}: {type(e).__name__}: {e}")
                drafts[sid] = {}
    print(f"drafted {sum(1 for d in drafts.values() if d)}/{len(rows)} docs")

    # Blind sampling is CLUSTER-aware: near-duplicate syllabi (draft field overlap
    # >= 0.8) go blind together — otherwise a blind doc's hidden draft leaks via
    # its assisted twin and the Rule-4 anchoring check is defeated.
    clusters = _cluster_by_draft_similarity(rows, drafts)
    rng = random.Random(args.seed)
    order = list(clusters)
    rng.shuffle(order)
    target = max(1, round(len(rows) * args.blind))
    blind_ids: set[str] = set()
    for cluster in order:
        if len(blind_ids) >= target:
            break
        blind_ids.update(cluster)
    multi = [c for c in clusters if len(c) > 1]
    print(f"{len(rows)} docs in {len(clusters)} clusters ({len(multi)} multi-doc); "
          f"blind subset ({len(blind_ids)}): {sorted(blind_ids)}")

    out_dir = Path("data/gold")
    out_dir.mkdir(parents=True, exist_ok=True)
    suffix = f"_{args.out_prefix}" if args.out_prefix else ""
    with open(out_dir / f"drafts{suffix}.jsonl", "w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps({"syllabus_id": r["syllabus_id"],
                                "doc_id": r.get("doc_id"),
                                "blind": r["syllabus_id"] in blind_ids,
                                "draft": drafts.get(r["syllabus_id"], {})}, ensure_ascii=False) + "\n")

    # --- workbook -----------------------------------------------------------
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    guide = wb.active
    guide.title = "안내"
    lines = [
        ("Gold 검수 — 순환 방지 5규칙 (v5 §3-2)", ""),
        ("Rule 1", "확정(Y) 표시된 셀만 gold. 초안은 시드일 뿐."),
        ("Rule 2", "판단 근거는 항상 '원문' 시트. 방법 출력(룰/LLM/하이브리드)은 이 파일에 없음."),
        ("Rule 3", "정답 칸은 비어 시작. 초안이 맞으면 복사, 틀리면 원문 기준으로 작성 — 편집률을 잽니다."),
        ("Rule 4", "초안 칸이 [BLIND]인 문서는 원문만 보고 처음부터 라벨링."),
        ("Rule 5", "날짜는 raw+날짜종류로만 (Week 8 → 'Week 8 | relative'). 실제 날짜 변환 금지."),
        ("", ""),
        ("표기 규칙", NOTATION),
    ]
    for i, (a, b) in enumerate(lines, 1):
        guide.cell(row=i, column=1, value=a).font = Font(bold=bool(a and not b) or a.startswith("Rule"))
        c = guide.cell(row=i, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    guide.column_dimensions["A"].width = 14
    guide.column_dimensions["B"].width = 110

    ws = wb.create_sheet("검수")
    header = ["syllabus_id", "field", "초안(참고용)", "정답(입력)", "확정(Y)", "메모"]
    ws.append(header)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="E8EAF0")
    blind_fill = PatternFill("solid", start_color="FFF3CD")
    for r in rows:
        sid = r["syllabus_id"]
        is_blind = sid in blind_ids
        for f in FIELDS:
            draft = None if is_blind else (drafts.get(sid, {}) or {}).get(f)
            ws.append([sid, f, "[BLIND]" if is_blind else (draft or ""), "", "", ""])
            for cell in ws[ws.max_row]:
                cell.number_format = "@"   # text — stop Excel coercing dates/times
                if is_blind:
                    cell.fill = blind_fill
    for col, w in zip("ABCDEF", (12, 12, 46, 46, 8, 24)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    src = wb.create_sheet("원문")
    src.append(["syllabus_id", "doc_id", "원문텍스트"])
    for r in rows:
        src.append([r["syllabus_id"], r.get("doc_id") or "", r["source_text"]])
        src.cell(row=src.max_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    src.column_dimensions["A"].width = 12
    src.column_dimensions["B"].width = 26
    src.column_dimensions["C"].width = 140
    src.freeze_panes = "A2"

    path = out_dir / f"gold_review{suffix}.xlsx"
    wb.save(path)
    print(f"wrote {path}  (검수 {len(rows)*len(FIELDS)} rows; 원문 {len(rows)}; drafts hidden for blind docs)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
